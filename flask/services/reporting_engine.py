"""
Enhanced Reporting Engine for Flask Migration
Provides comprehensive report generation with PDF/Excel output,
scheduled reports, and integration with analytics services.
"""

from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union
import uuid
import json
import os
from io import BytesIO
import base64

# Import data processing
import pandas as pd
import numpy as np

# Try to import advanced reporting libraries
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import LineChart, Reference, BarChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import analytics services for report content
from services.enhanced_correlation_service import enhanced_correlation_service, CorrelationConfig
from services.trend_analysis_service import trend_analysis_service, TrendConfig
from services.statistical_analysis_service import statistical_analysis_service, StatisticalConfig

from config.advanced_logging_config import get_advanced_logger

logger = get_advanced_logger(__name__)

class EnhancedReportingEngine:
    """Enhanced reporting engine with real PDF/Excel generation"""
    
    def __init__(self):
        self.logger = logger
        self.report_templates = {
            'water_quality_comprehensive': {
                'name': 'Comprehensive Water Quality Analysis',
                'description': 'Complete analysis including statistics, trends, and correlations',
                'parameters': ['sites', 'date_range', 'format', 'include_analytics'],
                'analytics': ['statistical', 'correlation', 'trend'],
                'formats': ['pdf', 'excel', 'csv']
            },
            'correlation_analysis_report': {
                'name': 'Parameter Correlation Analysis',
                'description': 'Detailed correlation analysis with significance testing',
                'parameters': ['sites', 'date_range', 'parameters', 'format'],
                'analytics': ['correlation'],
                'formats': ['pdf', 'excel']
            },
            'trend_forecast_report': {
                'name': 'Trend Analysis & Forecasting',
                'description': 'Time series analysis with forecasting and change detection',
                'parameters': ['sites', 'date_range', 'parameters', 'forecast_periods', 'format'],
                'analytics': ['trend'],
                'formats': ['pdf', 'excel']
            },
            'statistical_summary': {
                'name': 'Statistical Summary Report',
                'description': 'Comprehensive statistical analysis and data quality assessment',
                'parameters': ['sites', 'date_range', 'parameters', 'format'],
                'analytics': ['statistical'],
                'formats': ['pdf', 'excel', 'csv']
            },
            'site_comparison_advanced': {
                'name': 'Advanced Site Comparison',
                'description': 'Multi-dimensional comparison with statistical testing',
                'parameters': ['sites', 'date_range', 'metrics', 'format'],
                'analytics': ['statistical', 'correlation'],
                'formats': ['pdf', 'excel']
            }
        }
        
        self.report_history = []
        self.scheduled_reports = []
        
        # Create reports directory
        self.reports_dir = os.path.join(os.getcwd(), 'reports')
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_report(self, report_type: str, data: pd.DataFrame,
                       start_date: datetime, end_date: datetime, 
                       sites: List[str], format_type: str = 'pdf',
                       options: Dict[str, Any] = None) -> Dict[str, Any]:
        """Generate a comprehensive report with analytics"""
        
        try:
            self.logger.info(f"📊 Generating {report_type} report ({format_type})")
            
            if report_type not in self.report_templates:
                raise ValueError(f"Unknown report type: {report_type}")
            
            template = self.report_templates[report_type]
            if format_type not in template['formats']:
                raise ValueError(f"Format {format_type} not supported for {report_type}")
            
            options = options or {}
            report_id = str(uuid.uuid4())
            
            # Prepare data for analysis
            if data.empty:
                raise ValueError("No data available for report generation")
            
            # Generate analytics based on report type
            analytics_results = self._perform_analytics_for_report(
                data, template['analytics'], options
            )
            
            # Generate report file
            if format_type == 'pdf':
                file_path = self._generate_pdf_report(
                    report_id, report_type, data, analytics_results, 
                    start_date, end_date, sites, options
                )
            elif format_type == 'excel':
                file_path = self._generate_excel_report(
                    report_id, report_type, data, analytics_results,
                    start_date, end_date, sites, options
                )
            elif format_type == 'csv':
                file_path = self._generate_csv_report(
                    report_id, report_type, data, analytics_results,
                    start_date, end_date, sites, options
                )
            else:
                raise ValueError(f"Unsupported format: {format_type}")
            
            # Calculate file size
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            file_size_mb = file_size / (1024 * 1024)
            
            # Create report metadata
            report_metadata = {
                'report_id': report_id,
                'report_type': report_type,
                'report_name': template['name'],
                'format': format_type,
                'generated_at': datetime.now(),
                'start_date': start_date,
                'end_date': end_date,
                'sites': sites,
                'file_path': file_path,
                'file_size': file_size,
                'file_size_mb': f'{file_size_mb:.2f}MB',
                'status': 'completed',
                'analytics_included': template['analytics'],
                'options': options
            }
            
            # Add to history
            self.report_history.append(report_metadata)
            
            self.logger.info(f"✅ Report generated successfully: {report_id}")
            
            return {
                'success': True,
                'report_metadata': {
                    'report_id': report_id,
                    'report_name': template['name'],
                    'format': format_type,
                    'generated_at': report_metadata['generated_at'].isoformat(),
                    'file_size': report_metadata['file_size_mb'],
                    'download_url': f'/api/v1/reports/download/{report_id}',
                    'analytics_summary': self._create_analytics_summary(analytics_results)
                }
            }
            
        except Exception as e:
            self.logger.error(f"❌ Error generating report: {e}", exc_info=True)
            return {
                'success': False,
                'error': str(e),
                'report_id': None
            }
    
    def _perform_analytics_for_report(self, data: pd.DataFrame, 
                                    analytics_types: List[str],
                                    options: Dict[str, Any]) -> Dict[str, Any]:
        """Perform analytics based on report requirements"""
        
        analytics_results = {}
        
        # Get numeric parameters
        numeric_params = data.select_dtypes(include=[np.number]).columns.tolist()
        # Filter out timestamp/id columns
        numeric_params = [col for col in numeric_params if not any(
            keyword in col.lower() for keyword in ['id', 'timestamp', 'index']
        )]
        
        if not numeric_params:
            return analytics_results
        
        # Statistical analysis
        if 'statistical' in analytics_types:
            try:
                config = StatisticalConfig(
                    confidence_level=options.get('confidence_level', 0.95),
                    outlier_method=options.get('outlier_method', 'iqr')
                )
                stats_results = statistical_analysis_service.analyze_statistics(
                    data, numeric_params, config
                )
                analytics_results['statistical'] = stats_results
            except Exception as e:
                self.logger.warning(f"Statistical analysis failed: {e}")
        
        # Correlation analysis
        if 'correlation' in analytics_types:
            try:
                config = CorrelationConfig(
                    method=options.get('correlation_method', 'pearson'),
                    correlation_threshold=options.get('correlation_threshold', 0.3)
                )
                corr_results = enhanced_correlation_service.analyze_correlations(
                    data, config, numeric_params
                )
                analytics_results['correlation'] = corr_results
            except Exception as e:
                self.logger.warning(f"Correlation analysis failed: {e}")
        
        # Trend analysis
        if 'trend' in analytics_types:
            try:
                config = TrendConfig(
                    forecast_periods=options.get('forecast_periods', 24),
                    confidence_level=options.get('confidence_level', 0.95)
                )
                trend_results = trend_analysis_service.analyze_trends(
                    data, numeric_params, config
                )
                analytics_results['trend'] = trend_results
            except Exception as e:
                self.logger.warning(f"Trend analysis failed: {e}")
        
        return analytics_results
    
    def _generate_pdf_report(self, report_id: str, report_type: str,
                           data: pd.DataFrame, analytics: Dict[str, Any],
                           start_date: datetime, end_date: datetime,
                           sites: List[str], options: Dict[str, Any]) -> str:
        """Generate PDF report using ReportLab"""
        
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("ReportLab not available for PDF generation")
        
        filename = f"{report_type}_{report_id}.pdf"
        file_path = os.path.join(self.reports_dir, filename)
        
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        template = self.report_templates[report_type]
        story.append(Paragraph(template['name'], title_style))
        story.append(Spacer(1, 12))
        
        # Report metadata
        story.append(Paragraph("Report Information", styles['Heading2']))
        
        meta_data = [
            ['Report ID:', report_id],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Period:', f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"],
            ['Sites:', ', '.join(sites)],
            ['Data Points:', str(len(data))]
        ]
        
        meta_table = Table(meta_data)
        meta_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(meta_table)
        story.append(Spacer(1, 20))
        
        # Data summary
        story.append(Paragraph("Data Summary", styles['Heading2']))
        
        summary_stats = data.describe()
        if not summary_stats.empty:
            # Convert summary statistics to table
            summary_data = [['Parameter'] + summary_stats.index.tolist()]
            
            for col in summary_stats.columns:
                row = [col] + [f"{val:.3f}" if isinstance(val, (int, float)) else str(val) 
                             for val in summary_stats[col].values]
                summary_data.append(row)
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(PageBreak())
        
        # Analytics sections
        self._add_analytics_to_pdf(story, analytics, styles)
        
        # Build PDF
        doc.build(story)
        
        return file_path
    
    def _generate_excel_report(self, report_id: str, report_type: str,
                             data: pd.DataFrame, analytics: Dict[str, Any],
                             start_date: datetime, end_date: datetime,
                             sites: List[str], options: Dict[str, Any]) -> str:
        """Generate Excel report using openpyxl"""
        
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("OpenPyXL not available for Excel generation")
        
        filename = f"{report_type}_{report_id}.xlsx"
        file_path = os.path.join(self.reports_dir, filename)
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Data sheet
        ws_data = wb.create_sheet("Raw Data")
        
        # Write data to Excel
        for r_idx, row in enumerate(data.itertuples(), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        
        template = self.report_templates[report_type]
        ws_summary['A1'] = template['name']
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        # Report metadata
        ws_summary['A3'] = 'Report ID:'
        ws_summary['B3'] = report_id
        ws_summary['A4'] = 'Generated:'
        ws_summary['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_summary['A5'] = 'Period:'
        ws_summary['B5'] = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        ws_summary['A6'] = 'Sites:'
        ws_summary['B6'] = ', '.join(sites)
        ws_summary['A7'] = 'Data Points:'
        ws_summary['B7'] = len(data)
        
        # Add analytics sheets
        self._add_analytics_to_excel(wb, analytics)
        
        wb.save(file_path)
        return file_path
    
    def _generate_csv_report(self, report_id: str, report_type: str,
                           data: pd.DataFrame, analytics: Dict[str, Any],
                           start_date: datetime, end_date: datetime,
                           sites: List[str], options: Dict[str, Any]) -> str:
        """Generate CSV report"""
        
        filename = f"{report_type}_{report_id}.csv"
        file_path = os.path.join(self.reports_dir, filename)
        
        # For CSV, just export the data with basic metadata
        with open(file_path, 'w') as f:
            # Write metadata as comments
            template = self.report_templates[report_type]
            f.write(f"# {template['name']}\n")
            f.write(f"# Report ID: {report_id}\n")
            f.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}\n")
            f.write(f"# Sites: {', '.join(sites)}\n")
            f.write(f"# Data Points: {len(data)}\n#\n")
        
        # Append actual data
        data.to_csv(file_path, mode='a', index=False)
        
        return file_path
    
    def _add_analytics_to_pdf(self, story: List, analytics: Dict[str, Any], styles) -> None:
        """Add analytics sections to PDF"""
        
        # Statistical Analysis
        if 'statistical' in analytics:
            story.append(Paragraph("Statistical Analysis", styles['Heading2']))
            
            for param, result in analytics['statistical'].items():
                story.append(Paragraph(f"{param} Statistics", styles['Heading3']))
                
                stats = result.descriptive_stats
                if not isinstance(stats, dict) or 'error' in stats:
                    continue
                    
                stats_data = [
                    ['Metric', 'Value'],
                    ['Count', f"{stats.get('count', 0):,}"],
                    ['Mean', f"{stats.get('mean', 0):.3f}"],
                    ['Median', f"{stats.get('median', 0):.3f}"],
                    ['Std Dev', f"{stats.get('std', 0):.3f}"],
                    ['Min', f"{stats.get('min', 0):.3f}"],
                    ['Max', f"{stats.get('max', 0):.3f}"]
                ]
                
                stats_table = Table(stats_data)
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(stats_table)
                story.append(Spacer(1, 12))
                
                # Add insights
                if result.insights:
                    story.append(Paragraph("Key Insights:", styles['Heading4']))
                    for insight in result.insights[:3]:  # Limit to top 3
                        story.append(Paragraph(f"• {insight}", styles['Normal']))
                    story.append(Spacer(1, 12))
        
        # Correlation Analysis
        if 'correlation' in analytics:
            story.append(Paragraph("Correlation Analysis", styles['Heading2']))
            
            corr_result = analytics['correlation']
            if corr_result.significant_correlations:
                story.append(Paragraph("Significant Correlations", styles['Heading3']))
                
                corr_data = [['Parameter 1', 'Parameter 2', 'Correlation', 'P-value', 'Strength']]
                
                for corr in corr_result.significant_correlations[:10]:  # Top 10
                    corr_data.append([
                        corr['parameter1'],
                        corr['parameter2'],
                        f"{corr['correlation']:.3f}",
                        f"{corr['p_value']:.4f}",
                        corr['strength']
                    ])
                
                corr_table = Table(corr_data)
                corr_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(corr_table)
                story.append(Spacer(1, 12))
    
    def _add_analytics_to_excel(self, wb: Workbook, analytics: Dict[str, Any]) -> None:
        """Add analytics sheets to Excel workbook"""
        
        # Statistical Analysis
        if 'statistical' in analytics:
            ws_stats = wb.create_sheet("Statistical Analysis")
            
            row = 1
            ws_stats[f'A{row}'] = 'Statistical Analysis Results'
            ws_stats[f'A{row}'].font = Font(size=14, bold=True)
            row += 2
            
            for param, result in analytics['statistical'].items():
                ws_stats[f'A{row}'] = f"{param} Statistics"
                ws_stats[f'A{row}'].font = Font(bold=True)
                row += 1
                
                stats = result.descriptive_stats
                if isinstance(stats, dict) and 'error' not in stats:
                    for stat_name, stat_value in stats.items():
                        ws_stats[f'A{row}'] = stat_name
                        ws_stats[f'B{row}'] = stat_value
                        row += 1
                row += 1  # Add space between parameters
        
        # Correlation Analysis
        if 'correlation' in analytics:
            ws_corr = wb.create_sheet("Correlation Analysis")
            
            row = 1
            ws_corr[f'A{row}'] = 'Correlation Analysis Results'
            ws_corr[f'A{row}'].font = Font(size=14, bold=True)
            row += 2
            
            corr_result = analytics['correlation']
            if corr_result.significant_correlations:
                # Headers
                headers = ['Parameter 1', 'Parameter 2', 'Correlation', 'P-value', 'Strength']
                for col, header in enumerate(headers, 1):
                    cell = ws_corr.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                
                row += 1
                
                # Data
                for corr in corr_result.significant_correlations:
                    ws_corr[f'A{row}'] = corr['parameter1']
                    ws_corr[f'B{row}'] = corr['parameter2']
                    ws_corr[f'C{row}'] = corr['correlation']
                    ws_corr[f'D{row}'] = corr['p_value']
                    ws_corr[f'E{row}'] = corr['strength']
                    row += 1
    
    def _create_analytics_summary(self, analytics: Dict[str, Any]) -> Dict[str, Any]:
        """Create summary of analytics results"""
        
        summary = {}
        
        if 'statistical' in analytics:
            summary['statistical'] = {
                'parameters_analyzed': len(analytics['statistical']),
                'total_outliers': sum(
                    result.outlier_analysis.get('statistical_summary', {}).get('total_unique_outliers', 0)
                    for result in analytics['statistical'].values()
                )
            }
        
        if 'correlation' in analytics:
            corr_result = analytics['correlation']
            summary['correlation'] = {
                'significant_correlations': len(corr_result.significant_correlations),
                'strongest_correlation': (
                    corr_result.significant_correlations[0]['correlation']
                    if corr_result.significant_correlations else 0
                )
            }
        
        if 'trend' in analytics:
            trend_results = analytics['trend']
            summary['trend'] = {
                'parameters_analyzed': len(trend_results),
                'trending_parameters': sum(
                    1 for result in trend_results.values()
                    if result.trend_summary.get('statistically_significant', False)
                )
            }
        
        return summary
    
    def get_report_history(self, limit: int = 20) -> List[Dict[str, Any]]:
        """Get report generation history"""
        
        # Sort by generation date, most recent first
        sorted_history = sorted(
            self.report_history, 
            key=lambda x: x['generated_at'], 
            reverse=True
        )
        
        # Convert datetime objects for JSON serialization
        history_list = []
        for report in sorted_history[:limit]:
            report_dict = report.copy()
            report_dict['generated_at'] = report_dict['generated_at'].isoformat()
            report_dict['start_date'] = report_dict['start_date'].isoformat()
            report_dict['end_date'] = report_dict['end_date'].isoformat()
            history_list.append(report_dict)
            
        return history_list
    
    def get_scheduled_reports(self) -> List[Dict[str, Any]]:
        """Get scheduled reports"""
        return self.scheduled_reports
    
    def add_scheduled_report(self, report_config: Dict[str, Any]) -> str:
        """Add a scheduled report"""
        report_id = str(uuid.uuid4())
        report_config['id'] = report_id
        report_config['created_at'] = datetime.now()
        report_config['last_run'] = None
        self.scheduled_reports.append(report_config)
        return report_id
    
    def update_scheduled_report(self, report_id: str, config: Dict[str, Any]) -> bool:
        """Update scheduled report"""
        for report in self.scheduled_reports:
            if report['id'] == report_id:
                report.update(config)
                return True
        return False
    
    def delete_scheduled_report(self, report_id: str) -> bool:
        """Delete scheduled report"""
        for i, report in enumerate(self.scheduled_reports):
            if report['id'] == report_id:
                del self.scheduled_reports[i]
                return True
        return False
    
    def get_report_by_id(self, report_id: str) -> Optional[Dict[str, Any]]:
        """Get specific report by ID"""
        for report in self.report_history:
            if report['report_id'] == report_id:
                return report
        return None
    
    def get_available_templates(self) -> Dict[str, Dict[str, Any]]:
        """Get available report templates"""
        return self.report_templates


# Compatibility: Keep old class name for existing imports
class ReportingEngine(EnhancedReportingEngine):
    """Legacy compatibility class"""
    pass


# Global reporting engine instance (enhanced)
reporting_engine = EnhancedReportingEngine()